import psycopg
from psycopg.rows import dict_row
from psycopg_pool import AsyncConnectionPool
import os
import datetime
import time
from config import DB_PATH

PREDEFINED_VEHICLES = ['102', '103', '106', '107', '108', '109', '112', '115', '117', '122', '123', '477', '478', '480', '481', '482', '484', '485', '488', '491', '492', '493', '494', '497', '615', '617', '499', '489', '487', '124', '125', '126', '127', '9154', '9155', '9156', '9157', '9158', '9159', '361', '362', '364', '809', '810', '961']

# Global connection pool
db_pool = None
_user_cache: dict[int, tuple[float, dict | None]] = {}
USER_CACHE_TTL_SECONDS = 60


def _invalidate_user_cache(telegram_id: int):
    _user_cache.pop(telegram_id, None)


def format_datetime(value) -> str:
    """Format PostgreSQL datetime values and legacy text timestamps consistently."""
    if not value:
        return '—'
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d %H:%M')
    return str(value)[:16].replace('T', ' ')

async def _check_connection(conn):
    """Pool health check — dead connections are replaced automatically."""
    try:
        await conn.execute("SELECT 1")
    except Exception:
        raise

async def init_db():
    global db_pool
    db_url = os.environ.get('DATABASE_URL')
    if not db_url:
        raise RuntimeError("DATABASE_URL is not configured")
    db_pool = AsyncConnectionPool(
        db_url,
        kwargs={"row_factory": dict_row},
        min_size=1,
        max_size=10,
        max_idle=300,           # 5 daqiqadan keyin idle ulanishni yop
        max_lifetime=3600,      # 1 soatdan keyin ulanishni yangilash
        reconnect_timeout=30,   # qayta ulanishga 30s vaqt
        check=_check_connection,
        open=False,
    )
    await db_pool.open(wait=True, timeout=30)
    # Vehicle metadata is kept in PostgreSQL so newly imported drivers/models persist.
    async with db_pool.connection() as db:
        await db.execute(
            """CREATE TABLE IF NOT EXISTS bot_fsm (
                   storage_key TEXT PRIMARY KEY,
                   state TEXT,
                   data JSONB NOT NULL DEFAULT '{}'::jsonb,
                   updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
               )"""
        )
        await db.execute("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS driver_name TEXT")
        await db.execute("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS driver_phone TEXT")
        await db.execute("ALTER TABLE vehicles ADD COLUMN IF NOT EXISTS vehicle_model TEXT")
        # Hot-path indexes used by role menus and request workflow filters.
        await db.execute("CREATE INDEX IF NOT EXISTS idx_requests_status ON requests(status)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_requests_creator_status ON requests(created_by, status)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_requests_courier_status ON requests(courier_id, status)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_requests_vehicle ON requests(vehicle_name)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_request_items_request ON request_items(request_id)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_users_role_approved ON users(role, is_approved)")
        await db.commit()
    print("Connected to Neon DB successfully with connection pool")


async def close_db():
    """Close the shared database pool during application shutdown."""
    global db_pool
    if db_pool is not None:
        await db_pool.close()
        db_pool = None


async def get_vehicle_counts() -> dict:
    """SoZ va nosoz mashinalar sonini qaytaradi (kesh uchun)."""
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                "SELECT status, COUNT(*) as cnt FROM vehicles GROUP BY status"
            )
            rows = await cursor.fetchall()
            result = {'soz': 0, 'nosoz': 0, 'total': 0}
            for row in rows:
                if row['status'] == 'soz':
                    result['soz'] = row['cnt']
                elif row['status'] == 'nosoz':
                    result['nosoz'] = row['cnt']
                result['total'] += row['cnt']
            return result


async def add_user(telegram_id: int, full_name: str, phone: str, role: str):
    global db_pool
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        cursor = await db.execute('SELECT COUNT(*) FROM users')
        row = await cursor.fetchone()
        count = row['count'] if row else 0
        from config import ADMIN_ID
        is_approved = 0
        final_role = role
        if ADMIN_ID and telegram_id == ADMIN_ID:
            final_role = 'super_admin'
            is_approved = 1
        elif count == 0:
            final_role = 'super_admin'
            is_approved = 1
        await db.execute('\n            INSERT INTO users (telegram_id, full_name, phone, role, is_approved, created_at)\n            VALUES (%s, %s, %s, %s, %s, %s)\n            ON CONFLICT(telegram_id) DO UPDATE SET\n                full_name = excluded.full_name,\n                phone = excluded.phone,\n                role = excluded.role,\n                is_approved = excluded.is_approved\n        ', (telegram_id, full_name, phone, final_role, is_approved, now))
        await db.commit()
        _invalidate_user_cache(telegram_id)
        return (final_role, is_approved)

async def get_user(telegram_id: int):
    cached = _user_cache.get(telegram_id)
    now = time.monotonic()
    if cached and now - cached[0] < USER_CACHE_TTL_SECONDS:
        return cached[1]

    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM users WHERE telegram_id = %s', (telegram_id,))
            user = await cursor.fetchone()
            _user_cache[telegram_id] = (now, user)
            return user

async def get_pending_users():
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM users WHERE is_approved = 0')
            return await cursor.fetchall()

async def approve_user(telegram_id: int, role: str=None):
    global db_pool
    async with db_pool.connection() as db:
        if role:
            await db.execute('UPDATE users SET is_approved = 1, role = %s WHERE telegram_id = %s', (role, telegram_id))
        else:
            await db.execute('UPDATE users SET is_approved = 1 WHERE telegram_id = %s', (telegram_id,))
        await db.commit()
        _invalidate_user_cache(telegram_id)

async def reject_user(telegram_id: int):
    global db_pool
    async with db_pool.connection() as db:
        await db.execute('DELETE FROM users WHERE telegram_id = %s', (telegram_id,))
        await db.commit()
        _invalidate_user_cache(telegram_id)

async def delete_user(telegram_id: int):
    """Permanently remove an employee account from the bot."""
    global db_pool
    async with db_pool.connection() as db:
        await db.execute('DELETE FROM users WHERE telegram_id = %s', (telegram_id,))
        await db.commit()
        _invalidate_user_cache(telegram_id)

async def update_user_role(telegram_id: int, role: str):
    global db_pool
    async with db_pool.connection() as db:
        await db.execute('UPDATE users SET role = %s WHERE telegram_id = %s', (role, telegram_id))
        await db.commit()
        _invalidate_user_cache(telegram_id)

async def get_users_by_role(role: str):
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM users WHERE role = %s AND is_approved = 1', (role,))
            return await cursor.fetchall()


async def get_approved_users():
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                'SELECT * FROM users WHERE is_approved = 1 ORDER BY full_name'
            )
            return await cursor.fetchall()


async def search_users(query: str):
    pattern = f'%{query}%'
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT * FROM users
                   WHERE full_name ILIKE %s OR phone ILIKE %s
                   ORDER BY full_name LIMIT 10''',
                (pattern, pattern),
            )
            return await cursor.fetchall()

async def create_request(created_by: int, description: str, vehicle_name: str=None, old_part_photo: str=None, qty_used: int=None, qty_left: int=None, request_type: str='repair'):
    global db_pool
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        cursor = await db.execute("\n            INSERT INTO requests (created_by, description, status, created_at, updated_at, vehicle_name, old_part_photo, quantity_used, quantity_left, request_type)\n            VALUES (%s, %s, 'pending_approval', %s, %s, %s, %s, %s, %s, %s)\n            RETURNING id\n        ", (created_by, description, now, now, vehicle_name, old_part_photo, qty_used, qty_left, request_type))
        row = await cursor.fetchone()
        await db.commit()
        return row['id']

async def get_request(request_id: int):
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('\n            SELECT r.*, \n                   u.full_name as creator_name, u.phone as creator_phone,\n                   app.full_name as approver_name,\n                   wh.full_name as warehouseman_name,\n                   cour.full_name as courier_name\n            FROM requests r\n            LEFT JOIN users u ON r.created_by = u.telegram_id\n            LEFT JOIN users app ON r.approved_by = app.telegram_id\n            LEFT JOIN users wh ON r.warehouse_released_by = wh.telegram_id\n            LEFT JOIN users cour ON r.courier_id = cour.telegram_id\n            WHERE r.id = %s\n        ', (request_id,))
            return await cursor.fetchone()

async def update_request_status(request_id: int, status: str, updated_by_id: int, role: str):
    global db_pool
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        if role in ['manager', 'super_admin', 'observer']:
            if status in ['approved', 'pending_admin_approval']:
                await db.execute('\n                    UPDATE requests SET status = %s, approved_by = %s, updated_at = %s WHERE id = %s\n                ', (status, updated_by_id, now, request_id))
            else:
                await db.execute('\n                    UPDATE requests SET status = %s, updated_at = %s WHERE id = %s\n                ', (status, now, request_id))
        elif role == 'warehouseman':
            await db.execute('\n                UPDATE requests SET status = %s, warehouse_released_by = %s, updated_at = %s WHERE id = %s\n            ', (status, updated_by_id, now, request_id))
        elif role == 'courier':
            await db.execute('\n                UPDATE requests SET status = %s, courier_id = %s, updated_at = %s WHERE id = %s\n            ', (status, updated_by_id, now, request_id))
        else:
            await db.execute('\n                UPDATE requests SET status = %s, updated_at = %s WHERE id = %s\n            ', (status, now, request_id))
        await db.commit()

async def update_installed_part_photo(request_id: int, photo_id: str):
    global db_pool
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        await db.execute('\n            UPDATE requests SET installed_part_photo = %s, updated_at = %s WHERE id = %s\n        ', (photo_id, now, request_id))
        await db.commit()

async def get_broken_vehicles():
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("SELECT name FROM vehicles WHERE status = 'nosoz'")
            rows = await cursor.fetchall()
            return [r['name'] for r in rows]

async def get_healthy_vehicles():
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("SELECT name FROM vehicles WHERE status = 'soz'")
            rows = await cursor.fetchall()
            return [r['name'] for r in rows]

async def get_all_vehicles():
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("SELECT name FROM vehicles ORDER BY name")
            rows = await cursor.fetchall()
            return [row['name'] for row in rows]

async def update_vehicle_status(name: str, status: str, reason: str=None):
    global db_pool
    async with db_pool.connection() as db:
        await db.execute('UPDATE vehicles SET status = %s, reason = %s WHERE name = %s', (status, reason, name.strip()))
        await db.commit()

async def check_vehicle_active_requests(name: str) -> bool:
    global db_pool
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT COUNT(*) FROM requests \n            WHERE vehicle_name = %s AND status NOT IN ('completed', 'rejected')\n        ", (name.strip(),))
            row = await cursor.fetchone()
            return row['count'] > 0 if row else False

async def update_request_installation_details(request_id: int, photo_id: str, items_used_map: dict=None):
    global db_pool
    req = await get_request(request_id)
    creator_id = req['created_by'] if req else None
    now = datetime.datetime.now().isoformat()
    items = await get_request_items(request_id)
    if items_used_map is None:
        items_used_map = {str(item['id']): item['quantity_requested'] for item in items}
    total_qty_used = sum((int(v) for v in items_used_map.values()))
    total_qty_requested = sum((item['quantity_requested'] for item in items))
    total_qty_left = total_qty_requested - total_qty_used
    async with db_pool.connection() as db:
        await db.execute('\n            UPDATE requests \n            SET installed_part_photo = %s, quantity_used = %s, quantity_left = %s, updated_at = %s \n            WHERE id = %s\n        ', (photo_id, total_qty_used, total_qty_left, now, request_id))
        await db.commit()
    # New workflow records rasxod when the mechanic receives parts from warehouse.
    # Keep this fallback only for historical requests completed before that handover.
    if req and req['status'] != 'issued_to_mechanic':
        for item in items:
            used_qty = int(items_used_map.get(str(item['id']), item['quantity_requested']))
            if used_qty > 0:
                await add_or_update_inventory_item(item['item_name'], -used_qty)
                async with db_pool.connection() as db:
                    await db.execute("\n                        INSERT INTO stock_transactions (item_name, type, quantity, user_id, request_id, created_at)\n                        VALUES (%s, 'rasxod', %s, %s, %s, %s)\n                    ", (item['item_name'], used_qty, creator_id, request_id, now))
                    await db.commit()

async def update_request_details(request_id: int, description: str, vehicle_name: str, old_part_photo: str, qty_used: int=None, qty_left: int=None, request_type: str='repair'):
    global db_pool
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        await db.execute("\n            UPDATE requests \n            SET description = %s, vehicle_name = %s, old_part_photo = %s, quantity_used = %s, quantity_left = %s, request_type = %s, status = 'pending_approval', updated_at = %s\n            WHERE id = %s\n        ", (description, vehicle_name, old_part_photo, qty_used, qty_left, request_type, now, request_id))
        await db.commit()

async def update_request_item(request_id: int, item_name: str, quantity: int):
    async with db_pool.connection() as db:
        await db.execute('\n            UPDATE request_items \n            SET item_name = %s, quantity_requested = %s, quantity_missing = %s, quantity_available = 0\n            WHERE request_id = %s\n        ', (item_name, quantity, quantity, request_id))
        await db.commit()

async def get_requests_by_status(status: str):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('\n            SELECT r.*, u.full_name as creator_name \n            FROM requests r \n            JOIN users u ON r.created_by = u.telegram_id \n            WHERE r.status = %s\n            ORDER BY r.id ASC\n        ', (status,))
            return await cursor.fetchall()

async def get_all_requests():
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('\n            SELECT r.*, u.full_name as creator_name \n            FROM requests r \n            JOIN users u ON r.created_by = u.telegram_id \n            ORDER BY r.id ASC\n        ')
            return await cursor.fetchall()

async def get_requests_movement():
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('\n            SELECT * FROM (\n                SELECT r.*, \n                       u_creator.full_name as creator_name,\n                       u_approver.full_name as approver_name,\n                       u_wh.full_name as warehouseman_name,\n                       u_courier.full_name as courier_name\n                FROM requests r\n                LEFT JOIN users u_creator ON r.created_by = u_creator.telegram_id\n                LEFT JOIN users u_approver ON r.approved_by = u_approver.telegram_id\n                LEFT JOIN users u_wh ON r.warehouse_released_by = u_wh.telegram_id\n                LEFT JOIN users u_courier ON r.courier_id = u_courier.telegram_id\n                ORDER BY r.updated_at DESC\n                LIMIT 15\n            ) ORDER BY updated_at DESC\n        ')
            return await cursor.fetchall()

async def get_my_requests(telegram_id: int):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT * FROM requests \n            WHERE created_by = %s AND status NOT IN ('completed', 'rejected')\n            ORDER BY id ASC\n        ", (telegram_id,))
            return await cursor.fetchall()

async def get_inventory_item(name: str):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM inventory WHERE name = %s', (name.strip(),))
            return await cursor.fetchone()

async def get_all_inventory():
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM inventory ORDER BY name')
            return await cursor.fetchall()


async def get_open_requests(limit: int = 30):
    """Return every request that still requires work or monitoring."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   LEFT JOIN users u ON u.telegram_id = r.created_by
                   WHERE r.status NOT IN ('completed', 'rejected')
                   ORDER BY r.updated_at DESC
                   LIMIT %s""",
                (limit,),
            )
            return await cursor.fetchall()


async def get_completed_requests(limit: int = 30):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   LEFT JOIN users u ON u.telegram_id = r.created_by
                   WHERE r.status = 'completed'
                   ORDER BY r.updated_at DESC
                   LIMIT %s""",
                (limit,),
            )
            return await cursor.fetchall()


async def get_my_completed_requests(telegram_id: int, limit: int = 30):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT * FROM requests
                   WHERE created_by = %s AND status = 'completed'
                   ORDER BY updated_at DESC
                   LIMIT %s""",
                (telegram_id, limit),
            )
            return await cursor.fetchall()


async def get_user_request_counts(telegram_id: int):
    """Counts used as live badges in mechanic/brigadier menus."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT
                   COUNT(*) AS total,
                   COUNT(*) FILTER (WHERE status NOT IN ('completed', 'rejected')) AS unfinished,
                     COUNT(*) FILTER (WHERE status = 'completed') AS completed,
                     COUNT(*) FILTER (WHERE status = 'ready_for_installation') AS ready_for_pickup
                   FROM requests WHERE created_by = %s""",
                (telegram_id,),
            )
            row = await cursor.fetchone()
            return dict(row) if row else {'total': 0, 'unfinished': 0, 'completed': 0}


async def get_leadership_menu_counts():
    """Live count badges for Super Admin, Manager and Manager 2 menus."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT
                       (SELECT COUNT(*) FROM users WHERE is_approved = 0) AS pending_users,
                       (SELECT COUNT(*) FROM users WHERE is_approved = 1) AS approved_users,
                       COUNT(*) FILTER (
                           WHERE status IN ('pending_approval', 'pending_admin_approval')
                       ) AS pending_requests,
                       COUNT(*) FILTER (
                           WHERE status NOT IN ('completed', 'rejected')
                       ) AS open_requests,
                       COUNT(*) FILTER (WHERE status = 'completed') AS completed_requests,
                       COUNT(*) AS all_requests,
                       (SELECT COUNT(*) FROM inventory) AS inventory_items
                   FROM requests"""
            )
            row = await cursor.fetchone()
            return dict(row) if row else {
                'pending_users': 0, 'approved_users': 0, 'pending_requests': 0,
                'open_requests': 0, 'completed_requests': 0, 'all_requests': 0,
                'inventory_items': 0,
            }

async def get_all_my_requests(telegram_id: int):
    """Return all requests created by an employee, including completed history."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                "SELECT * FROM requests WHERE created_by = %s ORDER BY id DESC",
                (telegram_id,),
            )
            return await cursor.fetchall()


async def get_requests_ready_for_pickup(telegram_id: int):
    """Requests whose purchased parts are in the warehouse and await their creator."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                """SELECT * FROM requests
                   WHERE created_by = %s
                     AND status = 'ready_for_installation'
                     AND courier_id IS NOT NULL
                     AND warehouse_released_by IS NOT NULL
                   ORDER BY updated_at ASC""",
                (telegram_id,),
            )
            return await cursor.fetchall()


async def issue_request_to_creator(request_id: int, creator_id: int):
    """Issue request items from warehouse and register the rasxod transaction."""
    now = datetime.datetime.now().isoformat()
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                "SELECT id, created_by, status FROM requests WHERE id = %s FOR UPDATE",
                (request_id,),
            )
            request_row = await cursor.fetchone()
            if not request_row or request_row['created_by'] != creator_id:
                raise ValueError("Zayavka topilmadi yoki sizga tegishli emas.")
            if request_row['status'] != 'ready_for_installation':
                raise ValueError("Bu zayavka skladdan olish holatida emas yoki allaqachon olingan.")

            await cursor.execute(
                "SELECT item_name, quantity_requested FROM request_items WHERE request_id = %s",
                (request_id,),
            )
            items = await cursor.fetchall()
            if not items:
                raise ValueError("Zayavka mahsulotlari topilmadi.")

            # Validate all locked stock rows before subtracting anything.
            for item in items:
                await cursor.execute("SELECT quantity FROM inventory WHERE name = %s FOR UPDATE", (item['item_name'],))
                stock_row = await cursor.fetchone()
                if not stock_row or stock_row['quantity'] < item['quantity_requested']:
                    available = stock_row['quantity'] if stock_row else 0
                    raise ValueError(f"{item['item_name']} omborda yetarli emas (mavjud: {available}).")

            for item in items:
                quantity = item['quantity_requested']
                await cursor.execute(
                    "UPDATE inventory SET quantity = quantity - %s WHERE name = %s",
                    (quantity, item['item_name']),
                )
                await cursor.execute(
                    """INSERT INTO stock_transactions
                       (item_name, type, quantity, user_id, request_id, created_at)
                       VALUES (%s, 'rasxod', %s, %s, %s, %s)""",
                    (item['item_name'], quantity, creator_id, request_id, now),
                )

            await cursor.execute(
                "UPDATE requests SET status = 'issued_to_mechanic', updated_at = %s WHERE id = %s",
                (now, request_id),
            )
        await db.commit()


async def upsert_vehicle_metadata(
    name: str,
    driver_name: str,
    vehicle_model: str,
    driver_phone: str | None = None,
):
    """Create/update vehicle owner metadata without changing its current status."""
    async with db_pool.connection() as db:
        await db.execute(
            """INSERT INTO vehicles
                   (name, status, reason, driver_name, vehicle_model, driver_phone)
               VALUES (%s, 'soz', NULL, %s, %s, %s)
               ON CONFLICT (name) DO UPDATE SET
                   driver_name = EXCLUDED.driver_name,
                   vehicle_model = EXCLUDED.vehicle_model,
                   driver_phone = COALESCE(EXCLUDED.driver_phone, vehicles.driver_phone)""",
            (
                name.strip(),
                driver_name.strip(),
                vehicle_model.strip(),
                driver_phone.strip() if driver_phone else None,
            ),
        )
        await db.commit()


def analyze_inventory_by_rules(stock: list[dict]) -> dict:
    """Create a deterministic inventory audit without external services."""
    total_items = len(stock)
    total_quantity = sum(int(item.get('quantity') or 0) for item in stock)
    empty = [item for item in stock if int(item.get('quantity') or 0) == 0]
    critical = [item for item in stock if 0 < int(item.get('quantity') or 0) <= 10]
    sufficient = [item for item in stock if int(item.get('quantity') or 0) > 10]

    risks = [f"{item['name']}: тугаган" for item in empty]
    risks.extend(f"{item['name']}: {item['quantity']} дона қолган" for item in critical)

    recommendations = []
    if empty:
        recommendations.append(f"Тугаган {len(empty)} та маҳсулотни биринчи навбатда харид қилиш.")
    if critical:
        recommendations.append(f"Қолдиғи 10 донадан кам бўлган {len(critical)} та маҳсулот учун буюртма тайёрлаш.")
    if not empty and not critical:
        recommendations.append("Барча маҳсулотлар етарли; режали назоратни давом эттириш.")

    return {
        'summary': (
            f"Жами {total_items} хил маҳсулот, {total_quantity} дона. "
            f"Етарли: {len(sufficient)}, кам қолган: {len(critical)}, тугаган: {len(empty)}."
        ),
        'risks': risks,
        'recommendations': recommendations,
    }


async def export_inventory_to_excel():
    """Create a compact, shareable Excel snapshot of the current inventory."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    stock = await get_all_inventory()
    analysis = analyze_inventory_by_rules(stock)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Ombor qoldiqlari'
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells('A1:D1')
    worksheet['A1'] = 'MO BUTLASH — OMBOR QOLDIQLARI'
    worksheet['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    worksheet['A1'].fill = PatternFill('solid', fgColor='1F4E79')
    worksheet['A1'].alignment = Alignment(horizontal='center')
    worksheet.row_dimensions[1].height = 26

    headers = ['T/r', 'Toifa', 'Mahsulot nomi', 'Miqdori (dona)']
    widths = [8, 24, 42, 20]
    border_side = Side(style='thin', color='B4C6E7')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for column, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = worksheet.cell(row=2, column=column, value=header)
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        worksheet.column_dimensions[cell.column_letter].width = width

    for index, item in enumerate(stock, start=1):
        category = 'Tayyor mahsulot' if item.get('category') == 'tayyor' else 'Butlovchi mahsulot'
        values = [index, category, item['name'], item['quantity']]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=index + 2, column=column, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(
                horizontal='center' if column in (1, 4) else 'left',
                vertical='center',
                wrap_text=True,
            )
            cell.border = border
            if index % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='EAF2F8')

    worksheet.freeze_panes = 'A3'
    worksheet.auto_filter.ref = f'A2:D{max(2, len(stock) + 2)}'

    analysis_sheet = workbook.create_sheet('Qoidaviy tahlil')
    analysis_sheet.sheet_view.showGridLines = False
    analysis_sheet.merge_cells('A1:B1')
    analysis_sheet['A1'] = 'OMBOR ҚОЛДИҚЛАРИ — ҚОИДАВИЙ ТАҲЛИЛ'
    analysis_sheet['A1'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    analysis_sheet['A1'].fill = PatternFill('solid', fgColor='1F4E79')
    analysis_sheet['A1'].alignment = Alignment(horizontal='center')
    analysis_sheet.column_dimensions['A'].width = 22
    analysis_sheet.column_dimensions['B'].width = 95

    def add_analysis_row(row: int, label: str, value: str, color: str = 'FFFFFF'):
        label_cell = analysis_sheet.cell(row=row, column=1, value=label)
        value_cell = analysis_sheet.cell(row=row, column=2, value=value)
        for cell in (label_cell, value_cell):
            cell.font = Font(name='Calibri', size=10, bold=cell.column == 1)
            cell.fill = PatternFill('solid', fgColor=color)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border

    add_analysis_row(3, 'Қисқа хулоса', analysis['summary'])
    add_analysis_row(5, 'Хавфлар', '\n'.join(f"• {risk}" for risk in analysis['risks']) or 'Аниқ хавф топилмади.', 'FFF2CC')
    add_analysis_row(7, 'Тавсиялар', '\n'.join(f"• {item}" for item in analysis['recommendations']), 'E2F0D9')
    for row in range(3, 9):
        analysis_sheet.row_dimensions[row].height = 42

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    file_path = f'MO_Ombor_Qoldiqlari_{timestamp}.xlsx'
    workbook.save(file_path)
    return file_path

async def add_or_update_inventory_item(name: str, quantity_change: int, category: str='butlovchi'):
    async with db_pool.connection() as db:
        name_clean = name.strip()
        cursor = await db.execute('SELECT quantity FROM inventory WHERE name = %s', (name_clean,))
        row = await cursor.fetchone()
        if row:
            new_qty = max(0, row['quantity'] + quantity_change)
            await db.execute('UPDATE inventory SET quantity = %s WHERE name = %s', (new_qty, name_clean))
        else:
            await db.execute('INSERT INTO inventory (name, quantity, category) VALUES (%s, %s, %s)', (name_clean, max(0, quantity_change), category))
        await db.commit()

async def update_inventory_manually(name: str, new_qty: int, category: str, user_id: int):
    async with db_pool.connection() as db:
        name_clean = name.strip()
        cursor = await db.execute('SELECT quantity FROM inventory WHERE name = %s', (name_clean,))
        row = await cursor.fetchone()
        old_qty = row['quantity'] if row else 0
        if row:
            await db.execute(
                'UPDATE inventory SET quantity = %s, category = %s WHERE name = %s',
                (new_qty, category, name_clean),
            )
        else:
            await db.execute(
                'INSERT INTO inventory (name, quantity, category) VALUES (%s, %s, %s)',
                (name_clean, new_qty, category),
            )
        diff = new_qty - old_qty
        if diff != 0:
            tx_type = 'prixod' if diff > 0 else 'rasxod'
            now = datetime.datetime.now().isoformat()
            await db.execute('\n                INSERT INTO stock_transactions (item_name, type, quantity, user_id, created_at)\n                VALUES (%s, %s, %s, %s, %s)\n            ', (name_clean, tx_type, abs(diff), user_id, now))
        await db.commit()

async def add_request_item(request_id: int, item_name: str, quantity_requested: int, quantity_available: int, quantity_missing: int):
    async with db_pool.connection() as db:
        await db.execute('\n            INSERT INTO request_items (request_id, item_name, quantity_requested, quantity_available, quantity_missing)\n            VALUES (%s, %s, %s, %s, %s)\n        ', (request_id, item_name.strip(), quantity_requested, quantity_available, quantity_missing))
        await db.commit()

async def get_request_items(request_id: int):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM request_items WHERE request_id = %s', (request_id,))
            return await cursor.fetchall()


async def delete_request_items(request_id: int):
    async with db_pool.connection() as db:
        await db.execute('DELETE FROM request_items WHERE request_id = %s', (request_id,))
        await db.commit()


async def get_vehicle_overview(vehicle_name: str):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   JOIN users u ON r.created_by = u.telegram_id
                   WHERE r.vehicle_name = %s
                     AND r.status NOT IN ('completed', 'rejected')
                   ORDER BY r.id DESC''',
                (vehicle_name,),
            )
            active_requests = await cursor.fetchall()
            await cursor.execute(
                '''SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   JOIN users u ON r.created_by = u.telegram_id
                   WHERE r.vehicle_name = %s
                     AND r.status IN ('completed', 'rejected')
                   ORDER BY r.id DESC LIMIT 5''',
                (vehicle_name,),
            )
            history_requests = await cursor.fetchall()
            await cursor.execute(
                '''SELECT status, reason, driver_name, driver_phone, vehicle_model
                   FROM vehicles WHERE name = %s''',
                (vehicle_name,),
            )
            vehicle = await cursor.fetchone()
            return active_requests, history_requests, vehicle


async def get_courier_missing_items(courier_id: int):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT r.id AS request_id, r.vehicle_name, r.status,
                          ri.item_name, ri.quantity_missing, r.courier_id,
                          r.request_type, r.price
                   FROM request_items ri
                   JOIN requests r ON ri.request_id = r.id
                   WHERE r.courier_id = %s
                     AND r.status IN ('delivering', 'searching', 'purchased')
                     AND ri.quantity_missing > 0
                   ORDER BY r.id ASC''',
                (courier_id,)
            )
            return await cursor.fetchall()


async def get_courier_menu_counts(courier_id: int):
    """Return the live counters shown on a supplier's main menu."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT
                       COUNT(*) FILTER (WHERE status = 'approved' AND courier_id IS NULL) AS available,
                       COUNT(*) FILTER (WHERE courier_id = %s
                                          AND status IN ('delivering', 'searching', 'purchased')) AS active,
                       COUNT(*) FILTER (WHERE courier_id = %s
                                          AND status = 'waiting_receipt') AS awaiting_receipt,
                       COUNT(*) FILTER (
                           WHERE courier_id = %s
                             AND status IN ('delivering', 'searching', 'purchased')
                             AND EXISTS (
                                 SELECT 1 FROM request_items ri
                                 WHERE ri.request_id = requests.id
                                   AND ri.quantity_missing > 0
                             )
                       ) AS searching_items
                   FROM requests''',
                (courier_id, courier_id, courier_id),
            )
            row = await cursor.fetchone()
            return dict(row) if row else {
                'available': 0, 'active': 0, 'awaiting_receipt': 0, 'searching_items': 0
            }


async def claim_courier_request(request_id: int, courier_id: int) -> bool:
    """Atomically assign a request so two suppliers cannot take it together."""
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        cursor = await db.execute(
            '''UPDATE requests
               SET status = 'delivering', courier_id = %s, updated_at = %s
               WHERE id = %s AND status = 'approved' AND courier_id IS NULL
               RETURNING id''',
            (courier_id, now, request_id),
        )
        claimed = await cursor.fetchone()
        await db.commit()
        return claimed is not None


async def courier_owns_active_request(request_id: int, courier_id: int) -> bool:
    """Allow supplier actions only on that supplier's active request."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT 1 FROM requests
                   WHERE id = %s AND courier_id = %s
                     AND status IN ('delivering', 'searching', 'purchased')''',
                (request_id, courier_id),
            )
            return await cursor.fetchone() is not None


async def get_courier_day_summary(courier_id: int, day_pattern: str):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT r.id, r.vehicle_name, r.price, r.updated_at,
                          ri.item_name, ri.quantity_requested
                   FROM requests r
                   JOIN request_items ri ON ri.request_id = r.id
                   WHERE r.courier_id = %s
                     AND r.status IN ('waiting_receipt', 'ready_for_installation', 'completed')
                     AND CAST(r.updated_at AS TEXT) LIKE %s
                   ORDER BY r.updated_at ASC''',
                (courier_id, day_pattern),
            )
            delivered_rows = await cursor.fetchall()
            await cursor.execute(
                '''SELECT COUNT(*) AS count FROM requests
                   WHERE courier_id = %s
                     AND status IN ('delivering', 'searching', 'purchased')''',
                (courier_id,),
            )
            row = await cursor.fetchone()
            return delivered_rows, row['count'] if row else 0


async def get_courier_active_requests(courier_id: int):
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   JOIN users u ON r.created_by = u.telegram_id
                   WHERE r.courier_id = %s
                     AND r.status IN ('delivering', 'searching', 'purchased')
                   ORDER BY r.id ASC''',
                (courier_id,),
            )
            return await cursor.fetchall()


async def get_courier_waiting_receipts(courier_id: int):
    """Requests already handed to the warehouse, awaiting warehouse confirmation."""
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute(
                '''SELECT r.*, u.full_name AS creator_name
                   FROM requests r
                   JOIN users u ON r.created_by = u.telegram_id
                   WHERE r.courier_id = %s AND r.status = 'waiting_receipt'
                   ORDER BY r.updated_at ASC''',
                (courier_id,),
            )
            return await cursor.fetchall()

async def update_stock_on_receipt(request_id: int):
    req = await get_request(request_id)
    wh_id = req['warehouse_released_by'] if req else None
    now = datetime.datetime.now().isoformat()
    items = await get_request_items(request_id)
    for item in items:
        if item['quantity_missing'] > 0:
            await add_or_update_inventory_item(item['item_name'], item['quantity_missing'])
            async with db_pool.connection() as db:
                await db.execute("\n                    INSERT INTO stock_transactions (item_name, type, quantity, user_id, request_id, created_at)\n                    VALUES (%s, 'prixod', %s, %s, %s, %s)\n                ", (item['item_name'], item['quantity_missing'], wh_id, request_id, now))
                await db.commit()

def parse_with_regex_excel(text: str) -> list:
    import re
    items = []
    normalized = text.lower()
    units = ['ta', 'dona', 'shtuk', 'шт', 'шт.', 'd', 'x']
    pattern1 = '(\\d+)\\s*(?:ta|dona|shtuk|шт|шт\\.|d|x|\\*|-)?\\s+([^0-9,;\\n]+)'
    matches = re.findall(pattern1, normalized)
    if matches:
        for qty_str, name_str in matches:
            name = name_str.strip().strip(',.;- \t')
            if not name or name in units:
                continue
            qty = int(qty_str)
            items.append({'name': name.capitalize(), 'qty': qty})
    if not items:
        pattern2 = '([^0-9,;\\n]+)\\s+(\\d+)\\s*(?:ta|dona|shtuk|шт|d)?'
        matches2 = re.findall(pattern2, normalized)
        if matches2:
            for name_str, qty_str in matches2:
                name = name_str.strip().strip(',.;- \t')
                if not name or name in units:
                    continue
                qty = int(qty_str)
                items.append({'name': name.capitalize(), 'qty': qty})
    return items

async def export_requests_to_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    STATUS_LABELS_UZ = {'pending_approval': 'Тасдиқлаш кутилмоқда', 'pending_admin_approval': 'Админ кутилмоқда', 'approved': 'Тасдиқланган', 'delivering': 'Йўлда', 'searching': 'Қидирилмоқда', 'purchased': 'Сотиб олинди', 'waiting_receipt': 'Қабул кутилмоқда', 'completed': 'Якунланди', 'rejected': 'Рад этилди', 'ready_for_installation': 'Ўрнатишга тайёр', 'issued_to_mechanic': 'Складдан олинди'}
    COLOR_HEADER_BG = '1F4E79'
    COLOR_HEADER_FONT = 'FFFFFF'
    COLOR_TITLE_BG = '2E75B6'
    COLOR_ALT_ROW = 'DEEAF1'
    COLOR_BORDER = 'ADB9CA'
    COLOR_STATUS = {'pending_approval': 'FFF2CC', 'pending_admin_approval': 'FFE699', 'approved': 'E2EFDA', 'delivering': 'FCE4D6', 'searching': 'FFF2CC', 'purchased': 'FCE4D6', 'waiting_receipt': 'DDEBF7', 'completed': 'C6EFCE', 'rejected': 'F4CCCC'}

    def header_style(cell, text):
        cell.value = text
        cell.font = Font(name='Calibri', bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()

    def thin_border():
        side = Side(style='thin', color=COLOR_BORDER)
        return Border(left=side, right=side, top=side, bottom=side)

    def data_cell(cell, value, row_num, status=None, is_num=False):
        cell.value = value
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='center' if is_num else 'left', vertical='center', wrap_text=True)
        cell.border = thin_border()
        if status and status in COLOR_STATUS:
            cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_STATUS[status])
        elif row_num % 2 == 0:
            cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_ALT_ROW)
    ws1 = wb.active
    ws1.title = '📋 Zayavkalar'
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells('A1:N1')
    title_cell = ws1['A1']
    title_cell.value = 'MO BUTLASH — ZAYAVKALAR HISOBOTI'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color=COLOR_HEADER_FONT)
    title_cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 28
    headers_zayavka = [('T/r', 6), ('Yaratuvchi (Mexanik/Brigadir)', 28), ('Zayavka Tavsifi', 35), ('Yaratilgan Vaqt', 18), ('Holati', 22), ('Boshqaruvchi', 22), ("Ta'minotchi", 22), ('Skladchik', 22), ('Mahsulot Nomi', 28), ("So'ralgan (dona)", 16), ('Omborda Bor Edi', 16), ('Keltirildi / Yetishmagan', 20), ('Ishlatildi (dona)', 16), ('Omborda Qoldi (dona)', 18)]
    for col_idx, (hdr, width) in enumerate(headers_zayavka, start=1):
        cell = ws1.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 36
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('\n            SELECT r.id, creator.full_name as creator_name, r.description, r.created_at, r.status,\n                   manager.full_name as manager_name, courier.full_name as courier_name,\n                   warehouse.full_name as warehouseman_name,\n                   ri.item_name, ri.quantity_requested, ri.quantity_available, ri.quantity_missing,\n                   r.quantity_used, r.quantity_left\n            FROM requests r\n            LEFT JOIN users creator ON r.created_by = creator.telegram_id\n            LEFT JOIN users manager ON r.approved_by = manager.telegram_id\n            LEFT JOIN users courier ON r.courier_id = courier.telegram_id\n            LEFT JOIN users warehouse ON r.warehouse_released_by = warehouse.telegram_id\n            LEFT JOIN request_items ri ON r.id = ri.request_id\n            ORDER BY r.id ASC\n        ')
            rows = await cursor.fetchall()
    processed_rows = []
    for r in rows:
        item_name = r['item_name'] or ''
        parsed = parse_with_regex_excel(item_name)
        if len(parsed) > 1:
            for item in parsed:
                v_row = dict(r)
                v_row['item_name'] = item['name']
                v_row['quantity_requested'] = item['qty']
                v_row['quantity_missing'] = item['qty']
                v_row['quantity_available'] = 0
                v_row['quantity_used'] = item['qty']
                v_row['quantity_left'] = 0
                processed_rows.append(v_row)
        else:
            v_row = dict(r)
            if item_name:
                v_row['item_name'] = item_name.capitalize()
            processed_rows.append(v_row)
    grouped_requests = []
    current_req_id = None
    current_group = []
    for r in processed_rows:
        req_id = r['id']
        if req_id != current_req_id:
            if current_group:
                grouped_requests.append(current_group)
            current_group = [r]
            current_req_id = req_id
        else:
            current_group.append(r)
    if current_group:
        grouped_requests.append(current_group)
    column_widths = [width for _, width in headers_zayavka]
    current_row = 3
    for group in grouped_requests:
        num_items = len(group)
        start_row = current_row
        end_row = current_row + num_items - 1
        for offset, r in enumerate(group):
            row_idx = current_row + offset
            row_vals_item = [r['id'], r['creator_name'] or '—', r['description'] or '—', format_datetime(r['created_at']), STATUS_LABELS_UZ.get(r['status'], r['status'] or '—'), r['manager_name'] or '—', r['courier_name'] or '—', r['warehouseman_name'] or '—', r['item_name'] or '—', r['quantity_requested'] or 0, r['quantity_available'] or 0, r['quantity_missing'] or 0, r['quantity_used'] if r['quantity_used'] is not None else '—', r['quantity_left'] if r['quantity_left'] is not None else '—']
            status_key = r['status']
            for col_idx, val in enumerate(row_vals_item, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                is_num = col_idx in (1, 10, 11, 12, 13, 14)
                data_cell(cell, val, row_idx, status_key, is_num)
            max_lines = 1
            for val, width in zip(row_vals_item, column_widths):
                if val is not None and isinstance(val, str):
                    lines = val.count('\n') + 1
                    for part in val.split('\n'):
                        part_len = len(part)
                        effective_width = max(5, width - 2)
                        if part_len > effective_width:
                            lines += (part_len - 1) // effective_width
                    max_lines = max(max_lines, lines)
            ws1.row_dimensions[row_idx].height = max(20, max_lines * 15 + 5)
        if num_items > 1:
            cols_to_merge = [1, 2, 3, 4, 5, 6, 7, 8, 13, 14]
            for col_idx in cols_to_merge:
                ws1.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                for r_idx in range(start_row, end_row + 1):
                    cell = ws1.cell(row=r_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal='center' if col_idx in (1, 13, 14) else 'left', vertical='center', wrap_text=True)
        current_row += num_items
    ws1.freeze_panes = 'A3'
    ws2 = wb.create_sheet(title='📦 Ombor Qoldiqlari')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:C1')
    title2 = ws2['A1']
    title2.value = 'OMBOR ZAXIRASI — JORIY QOLDIQLAR'
    title2.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title2.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26
    inv_headers = [('T/r', 6), ('Mahsulot Nomi', 35), ('Omborda Bor Miqdor (Dona)', 26), ('Turi', 22)]
    for col_idx, (hdr, width) in enumerate(inv_headers, start=1):
        cell = ws2.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT name, quantity, category FROM inventory ORDER BY name')
            inv_items = await cursor.fetchall()
    inventory_summary = {}
    for item in inv_items:
        name = item['name'] or ''
        parsed = parse_with_regex_excel(name)
        category = item['category']
        if len(parsed) > 1:
            for parsed_item in parsed:
                p_name = parsed_item['name']
                p_qty = parsed_item['qty']
                if p_name in inventory_summary:
                    inventory_summary[p_name]['quantity'] += p_qty
                else:
                    inventory_summary[p_name] = {'quantity': p_qty, 'category': category}
        else:
            p_name = name.capitalize()
            p_qty = item['quantity']
            if p_name in inventory_summary:
                inventory_summary[p_name]['quantity'] += p_qty
            else:
                inventory_summary[p_name] = {'quantity': p_qty, 'category': category}
    processed_inv_items = []
    for p_name, details in sorted(inventory_summary.items()):
        processed_inv_items.append({'name': p_name, 'quantity': details['quantity'], 'category': details['category']})
    for inv_idx, item in enumerate(processed_inv_items, start=3):
        qty = item['quantity']
        bg = 'C6EFCE' if qty > 0 else 'F4CCCC'
        category_text = 'Tayyor mahsulot' if item['category'] == 'tayyor' else 'Butlovchi mahsulot'
        num_cell = ws2.cell(row=inv_idx, column=1, value=inv_idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws2.cell(row=inv_idx, column=2, value=item['name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws2.cell(row=inv_idx, column=3, value=qty)
        qty_cell.font = Font(name='Calibri', size=11, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        cat_cell = ws2.cell(row=inv_idx, column=4, value=category_text)
        cat_cell.font = Font(name='Calibri', size=10)
        cat_cell.alignment = Alignment(horizontal='center', vertical='center')
        cat_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        cat_cell.border = thin_border()
        ws2.row_dimensions[inv_idx].height = 20
    ws2.freeze_panes = 'A3'
    ws3 = wb.create_sheet(title='📥 Kirimlar (Prixod)')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:F1')
    title3 = ws3['A1']
    title3.value = 'OMBORGA KIRIM QILINGAN TOVARLAR (PRIXOD)'
    title3.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title3.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26
    tx_headers = [('T/r', 6), ('Mahsulot Nomi', 35), ('Miqdori (dona)', 18), ("Mas'ul Xodim", 28), ('Zayavka ID', 15), ('Sana / Vaqt', 22)]
    for col_idx, (hdr, width) in enumerate(tx_headers, start=1):
        cell = ws3.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT t.item_name, t.quantity, u.full_name as user_name, t.request_id, t.created_at\n            FROM stock_transactions t\n            LEFT JOIN users u ON t.user_id = u.telegram_id\n            WHERE t.type = 'prixod'\n            ORDER BY t.id DESC\n        ")
            prixod_rows = await cursor.fetchall()
    processed_prixod_rows = []
    for tx in prixod_rows:
        name = tx['item_name'] or ''
        parsed = parse_with_regex_excel(name)
        if len(parsed) > 1:
            for parsed_item in parsed:
                v_tx = dict(tx)
                v_tx['item_name'] = parsed_item['name']
                v_tx['quantity'] = parsed_item['qty']
                processed_prixod_rows.append(v_tx)
        else:
            v_tx = dict(tx)
            if name:
                v_tx['item_name'] = name.capitalize()
            processed_prixod_rows.append(v_tx)
    for idx, r in enumerate(processed_prixod_rows, start=3):
        bg = 'E2EFDA'
        num_cell = ws3.cell(row=idx, column=1, value=idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws3.cell(row=idx, column=2, value=r['item_name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws3.cell(row=idx, column=3, value=r['quantity'])
        qty_cell.font = Font(name='Calibri', size=10, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        user_cell = ws3.cell(row=idx, column=4, value=r['user_name'] or 'Avtomatik / Tizim')
        user_cell.font = Font(name='Calibri', size=10)
        user_cell.alignment = Alignment(horizontal='left', vertical='center')
        user_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        user_cell.border = thin_border()
        req_cell = ws3.cell(row=idx, column=5, value=r['request_id'] if r['request_id'] else '—')
        req_cell.font = Font(name='Calibri', size=10)
        req_cell.alignment = Alignment(horizontal='center', vertical='center')
        req_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        req_cell.border = thin_border()
        date_cell = ws3.cell(row=idx, column=6, value=format_datetime(r['created_at']))
        date_cell.font = Font(name='Calibri', size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        date_cell.border = thin_border()
        ws3.row_dimensions[idx].height = 20
    ws3.freeze_panes = 'A3'
    ws4 = wb.create_sheet(title='📤 Chiqimlar (Rasxod)')
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells('A1:F1')
    title4 = ws4['A1']
    title4.value = 'OMBORDAN CHIQARILGAN TOVARLAR (RASXOD)'
    title4.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title4.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title4.alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 26
    for col_idx, (hdr, width) in enumerate(tx_headers, start=1):
        cell = ws4.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws4.column_dimensions[get_column_letter(col_idx)].width = width
    ws4.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT t.item_name, t.quantity, u.full_name as user_name, t.request_id, t.created_at\n            FROM stock_transactions t\n            LEFT JOIN users u ON t.user_id = u.telegram_id\n            WHERE t.type = 'rasxod'\n            ORDER BY t.id DESC\n        ")
            rasxod_rows = await cursor.fetchall()
    processed_rasxod_rows = []
    for tx in rasxod_rows:
        name = tx['item_name'] or ''
        parsed = parse_with_regex_excel(name)
        if len(parsed) > 1:
            for parsed_item in parsed:
                v_tx = dict(tx)
                v_tx['item_name'] = parsed_item['name']
                v_tx['quantity'] = parsed_item['qty']
                processed_rasxod_rows.append(v_tx)
        else:
            v_tx = dict(tx)
            if name:
                v_tx['item_name'] = name.capitalize()
            processed_rasxod_rows.append(v_tx)
    for idx, r in enumerate(processed_rasxod_rows, start=3):
        bg = 'FCE4D6'
        num_cell = ws4.cell(row=idx, column=1, value=idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws4.cell(row=idx, column=2, value=r['item_name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws4.cell(row=idx, column=3, value=r['quantity'])
        qty_cell.font = Font(name='Calibri', size=10, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        user_cell = ws4.cell(row=idx, column=4, value=r['user_name'] or 'Avtomatik / Tizim')
        user_cell.font = Font(name='Calibri', size=10)
        user_cell.alignment = Alignment(horizontal='left', vertical='center')
        user_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        user_cell.border = thin_border()
        req_cell = ws4.cell(row=idx, column=5, value=r['request_id'] if r['request_id'] else '—')
        req_cell.font = Font(name='Calibri', size=10)
        req_cell.alignment = Alignment(horizontal='center', vertical='center')
        req_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        req_cell.border = thin_border()
        date_cell = ws4.cell(row=idx, column=6, value=format_datetime(r['created_at']))
        date_cell.font = Font(name='Calibri', size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        date_cell.border = thin_border()
        ws4.row_dimensions[idx].height = 20
    ws4.freeze_panes = 'A3'
    import datetime
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    file_path = f'MO_Butlash_Hisobot_{timestamp}.xlsx'
    wb.save(file_path)
    return file_path

async def export_daily_report_to_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime
    wb = openpyxl.Workbook()
    COLOR_HEADER_BG = '1F4E79'
    COLOR_HEADER_FONT = 'FFFFFF'
    COLOR_TITLE_BG = '2E75B6'
    COLOR_ALT_ROW = 'DEEAF1'
    COLOR_BORDER = 'ADB9CA'

    def header_style(cell, text):
        cell.value = text
        cell.font = Font(name='Calibri', bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()

    def thin_border():
        side = Side(style='thin', color=COLOR_BORDER)
        return Border(left=side, right=side, top=side, bottom=side)
    today_str = datetime.date.today().isoformat()
    ws1 = wb.active
    ws1.title = '📦 Ombor Qoldiqlari'
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells('A1:D1')
    title1 = ws1['A1']
    title1.value = f'OMBOR QOLDIQLARI MONITORI ({today_str})'
    title1.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title1.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title1.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 26
    inv_headers = [('T/r', 6), ('Mahsulot Nomi', 35), ('Omborda Bor Miqdor (Dona)', 26), ('Turi', 22)]
    for col_idx, (hdr, width) in enumerate(inv_headers, start=1):
        cell = ws1.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width
    ws1.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT name, quantity, category FROM inventory ORDER BY name')
            inv_items = await cursor.fetchall()
    inventory_summary = {}
    for item in inv_items:
        name = item['name'] or ''
        parsed = parse_with_regex_excel(name)
        category = item['category']
        if len(parsed) > 1:
            for parsed_item in parsed:
                p_name = parsed_item['name']
                p_qty = parsed_item['qty']
                if p_name in inventory_summary:
                    inventory_summary[p_name]['quantity'] += p_qty
                else:
                    inventory_summary[p_name] = {'quantity': p_qty, 'category': category}
        else:
            p_name = name.capitalize()
            p_qty = item['quantity']
            if p_name in inventory_summary:
                inventory_summary[p_name]['quantity'] += p_qty
            else:
                inventory_summary[p_name] = {'quantity': p_qty, 'category': category}
    processed_inv_items = []
    for p_name, details in sorted(inventory_summary.items()):
        processed_inv_items.append({'name': p_name, 'quantity': details['quantity'], 'category': details['category']})
    for inv_idx, item in enumerate(processed_inv_items, start=3):
        qty = item['quantity']
        bg = 'C6EFCE' if qty > 0 else 'F4CCCC'
        category_text = 'Tayyor mahsulot' if item['category'] == 'tayyor' else 'Butlovchi mahsulot'
        num_cell = ws1.cell(row=inv_idx, column=1, value=inv_idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws1.cell(row=inv_idx, column=2, value=item['name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws1.cell(row=inv_idx, column=3, value=qty)
        qty_cell.font = Font(name='Calibri', size=11, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        cat_cell = ws1.cell(row=inv_idx, column=4, value=category_text)
        cat_cell.font = Font(name='Calibri', size=10)
        cat_cell.alignment = Alignment(horizontal='center', vertical='center')
        cat_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        cat_cell.border = thin_border()
        ws1.row_dimensions[inv_idx].height = 20
    ws1.freeze_panes = 'A3'
    ws2 = wb.create_sheet(title='📥 Bugungi Kirimlar (Prixod)')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:F1')
    title2 = ws2['A1']
    title2.value = f'BUGUNGI KIRIM QILINGAN TOVARLAR ({today_str})'
    title2.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title2.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26
    tx_headers = [('T/r', 6), ('Mahsulot Nomi', 35), ('Miqdori (dona)', 18), ("Mas'ul Xodim", 28), ('Zayavka ID', 15), ('Sana / Vaqt', 22)]
    for col_idx, (hdr, width) in enumerate(tx_headers, start=1):
        cell = ws2.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT t.item_name, t.quantity, u.full_name as user_name, t.request_id, t.created_at\n            FROM stock_transactions t\n            LEFT JOIN users u ON t.user_id = u.telegram_id\n            WHERE t.type = 'prixod' AND CAST(t.created_at AS TEXT) LIKE %s\n            ORDER BY t.id DESC\n        ", (f'{today_str}%',))
            prixod_rows = await cursor.fetchall()
    processed_prixod_rows = []
    for tx in prixod_rows:
        name = tx['item_name'] or ''
        parsed = parse_with_regex_excel(name)
        if len(parsed) > 1:
            for parsed_item in parsed:
                v_tx = dict(tx)
                v_tx['item_name'] = parsed_item['name']
                v_tx['quantity'] = parsed_item['qty']
                processed_prixod_rows.append(v_tx)
        else:
            v_tx = dict(tx)
            if name:
                v_tx['item_name'] = name.capitalize()
            processed_prixod_rows.append(v_tx)
    for idx, r in enumerate(processed_prixod_rows, start=3):
        bg = 'E2EFDA'
        num_cell = ws2.cell(row=idx, column=1, value=idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws2.cell(row=idx, column=2, value=r['item_name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws2.cell(row=idx, column=3, value=r['quantity'])
        qty_cell.font = Font(name='Calibri', size=10, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        user_cell = ws2.cell(row=idx, column=4, value=r['user_name'] or 'Avtomatik / Tizim')
        user_cell.font = Font(name='Calibri', size=10)
        user_cell.alignment = Alignment(horizontal='left', vertical='center')
        user_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        user_cell.border = thin_border()
        req_cell = ws2.cell(row=idx, column=5, value=r['request_id'] if r['request_id'] else '—')
        req_cell.font = Font(name='Calibri', size=10)
        req_cell.alignment = Alignment(horizontal='center', vertical='center')
        req_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        req_cell.border = thin_border()
        date_cell = ws2.cell(row=idx, column=6, value=format_datetime(r['created_at']))
        date_cell.font = Font(name='Calibri', size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        date_cell.border = thin_border()
        ws2.row_dimensions[idx].height = 20
    ws2.freeze_panes = 'A3'
    ws3 = wb.create_sheet(title='📤 Bugungi Chiqimlar (Rasxod)')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:F1')
    title3 = ws3['A1']
    title3.value = f'BUGUNGI CHIQARILGAN TOVARLAR ({today_str})'
    title3.font = Font(name='Calibri', bold=True, size=13, color=COLOR_HEADER_FONT)
    title3.fill = PatternFill(fill_type='solid', fgColor=COLOR_TITLE_BG)
    title3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 26
    for col_idx, (hdr, width) in enumerate(tx_headers, start=1):
        cell = ws3.cell(row=2, column=col_idx)
        header_style(cell, hdr)
        ws3.column_dimensions[get_column_letter(col_idx)].width = width
    ws3.row_dimensions[2].height = 34
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute("\n            SELECT t.item_name, t.quantity, u.full_name as user_name, t.request_id, t.created_at\n            FROM stock_transactions t\n            LEFT JOIN users u ON t.user_id = u.telegram_id\n            WHERE t.type = 'rasxod' AND CAST(t.created_at AS TEXT) LIKE %s\n            ORDER BY t.id DESC\n        ", (f'{today_str}%',))
            rasxod_rows = await cursor.fetchall()
    processed_rasxod_rows = []
    for tx in rasxod_rows:
        name = tx['item_name'] or ''
        parsed = parse_with_regex_excel(name)
        if len(parsed) > 1:
            for parsed_item in parsed:
                v_tx = dict(tx)
                v_tx['item_name'] = parsed_item['name']
                v_tx['quantity'] = parsed_item['qty']
                processed_rasxod_rows.append(v_tx)
        else:
            v_tx = dict(tx)
            if name:
                v_tx['item_name'] = name.capitalize()
            processed_rasxod_rows.append(v_tx)
    for idx, r in enumerate(processed_rasxod_rows, start=3):
        bg = 'FCE4D6'
        num_cell = ws3.cell(row=idx, column=1, value=idx - 2)
        num_cell.font = Font(name='Calibri', size=10, bold=True)
        num_cell.alignment = Alignment(horizontal='center', vertical='center')
        num_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        num_cell.border = thin_border()
        name_cell = ws3.cell(row=idx, column=2, value=r['item_name'])
        name_cell.font = Font(name='Calibri', size=10)
        name_cell.alignment = Alignment(horizontal='left', vertical='center')
        name_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        name_cell.border = thin_border()
        qty_cell = ws3.cell(row=idx, column=3, value=r['quantity'])
        qty_cell.font = Font(name='Calibri', size=10, bold=True)
        qty_cell.alignment = Alignment(horizontal='center', vertical='center')
        qty_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        qty_cell.border = thin_border()
        user_cell = ws3.cell(row=idx, column=4, value=r['user_name'] or 'Avtomatik / Tizim')
        user_cell.font = Font(name='Calibri', size=10)
        user_cell.alignment = Alignment(horizontal='left', vertical='center')
        user_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        user_cell.border = thin_border()
        req_cell = ws3.cell(row=idx, column=5, value=r['request_id'] if r['request_id'] else '—')
        req_cell.font = Font(name='Calibri', size=10)
        req_cell.alignment = Alignment(horizontal='center', vertical='center')
        req_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        req_cell.border = thin_border()
        date_cell = ws3.cell(row=idx, column=6, value=format_datetime(r['created_at']))
        date_cell.font = Font(name='Calibri', size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        date_cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        date_cell.border = thin_border()
        ws3.row_dimensions[idx].height = 20
    ws3.freeze_panes = 'A3'
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    file_path = f'MO_Kunlik_Hisobot_{timestamp}.xlsx'
    wb.save(file_path)
    return file_path

async def split_request(original_request_id: int, missing_item_ids: list) -> int:
    """
    Original zayavkadagi topilmagan mahsulotlarni ajratib,
    yangi faol (approved) zayavka yaratadi.
    """
    async with db_pool.connection() as db:
        async with db.cursor() as cursor:
            await cursor.execute('SELECT * FROM requests WHERE id = %s', (original_request_id,))
            orig = await cursor.fetchone()
            if not orig:
                raise ValueError('Original request not found')
        now = datetime.datetime.now().isoformat()
        cursor = await db.execute("\n            INSERT INTO requests (\n                created_by, description, status, approved_by, \n                created_at, updated_at, vehicle_name, old_part_photo, request_type\n            )\n            VALUES (%s, %s, 'approved', %s, %s, %s, %s, %s, %s)\n            RETURNING id\n        ", (orig['created_by'], orig['description'], orig['approved_by'], orig['created_at'], now, orig['vehicle_name'], orig['old_part_photo'], orig['request_type']))
        new_request_id = (await cursor.fetchone())['id']
        for item_id in missing_item_ids:
            await db.execute('\n                UPDATE request_items \n                SET request_id = %s \n                WHERE id = %s AND request_id = %s\n            ', (new_request_id, item_id, original_request_id))
        await db.commit()
        return new_request_id

async def update_request_price(request_id: int, price: int):
    async with db_pool.connection() as db:
        now = datetime.datetime.now().isoformat()
        await db.execute('\n            UPDATE requests \n            SET price = %s, updated_at = %s\n            WHERE id = %s\n        ', (price, now, request_id))
        await db.commit()
