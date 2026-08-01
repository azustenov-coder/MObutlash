import asyncio
import os
import sys
import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

env_path = os.path.join(os.path.dirname(__file__), '.env')
load_dotenv(dotenv_path=env_path)

import database as db

async def export_to_excel():
    output_filename = "Avtomashinalar_Ruyxati.xlsx"
    file_path = os.path.join(os.path.dirname(__file__), output_filename)
    
    print("Database ulanmoqda...")
    await db.init_db()
    
    async with db.db_pool.connection() as conn:
        async with conn.cursor() as cur:
            # Fetch all vehicles with request stats
            await cur.execute("""
                SELECT 
                    v.name, 
                    v.vehicle_model, 
                    v.driver_name, 
                    v.driver_phone, 
                    v.status, 
                    v.reason,
                    COUNT(CASE WHEN r.status NOT IN ('completed', 'rejected') THEN 1 END) as active_reqs,
                    COUNT(CASE WHEN r.status = 'completed' THEN 1 END) as completed_reqs,
                    COUNT(r.id) as total_reqs
                FROM vehicles v
                LEFT JOIN requests r ON v.name = r.vehicle_name
                GROUP BY v.name, v.vehicle_model, v.driver_name, v.driver_phone, v.status, v.reason
                ORDER BY 
                    CASE WHEN v.status = 'nosoz' THEN 1 ELSE 2 END,
                    v.name;
            """)
            vehicles = await cur.fetchall()
            
    await db.close_db()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avtomashinalar"
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="595959")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    soz_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    soz_font = Font(name="Calibri", size=11, bold=True, color="375623")
    
    nosoz_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    nosoz_font = Font(name="Calibri", size=11, bold=True, color="C65911")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Title row
    ws.merge_cells("A1:J1")
    ws["A1"] = "MO BUTLASH - AVTOMASHINALAR RO'YXATI VA HOLATI"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Hujjat yaratilgan sana: {now_str} | Jami avtomashinalar: {len(vehicles)} ta"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    
    # Headers
    headers = [
        "№", 
        "Mashina Nomi / Raqami", 
        "Modeli", 
        "Haydovchi Ismi", 
        "Haydovchi Telefoni", 
        "Holati", 
        "Sabab / Faol Zayavkalar", 
        "Faol Zayavkalar", 
        "Bajarilgan Zayavkalar", 
        "Jami Zayavkalar"
    ]
    
    header_row = 4
    ws.row_dimensions[header_row].height = 26
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Data rows
    start_data_row = 5
    for idx, v in enumerate(vehicles, 1):
        current_row = start_data_row + idx - 1
        ws.row_dimensions[current_row].height = 24
        
        status_val = (v['status'] or 'soz').lower()
        status_display = "NOSOZ" if status_val == 'nosoz' else "SOZ"
        
        row_data = [
            idx,
            v['name'],
            v.get('vehicle_model') or '—',
            v.get('driver_name') or '—',
            v.get('driver_phone') or '—',
            status_display,
            v.get('reason') or '—',
            v.get('active_reqs', 0),
            v.get('completed_reqs', 0),
            v.get('total_reqs', 0)
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 5, 6, 8, 9, 10]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
            # Status specific styling
            if col_idx == 6:
                if status_display == "NOSOZ":
                    cell.fill = nosoz_fill
                    cell.font = nosoz_font
                else:
                    cell.fill = soz_fill
                    cell.font = soz_font
                    
    # Adjust column widths
    ws.column_dimensions['A'].width = 6    # №
    ws.column_dimensions['B'].width = 24   # Name
    ws.column_dimensions['C'].width = 18   # Model
    ws.column_dimensions['D'].width = 25   # Driver
    ws.column_dimensions['E'].width = 20   # Phone
    ws.column_dimensions['F'].width = 14   # Status
    ws.column_dimensions['G'].width = 50   # Reason
    ws.column_dimensions['H'].width = 16   # Active
    ws.column_dimensions['I'].width = 20   # Completed
    ws.column_dimensions['J'].width = 16   # Total
    
    wb.save(file_path)
    print(f"Excel fayl muvaffaqiyatli yaratildi: {file_path}")
    return file_path

if __name__ == "__main__":
    asyncio.run(export_to_excel())
